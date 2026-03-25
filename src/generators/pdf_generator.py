"""
BidGenie AI - PDF Proposal Generator
Generates professional, branded PDF bid proposals for Ace Plumbing.
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# ─── COLORS & STYLING ────────────────────────────────────────────────────────
ACE_BLUE = (0, 71, 133)        # Deep blue
ACE_LIGHT_BLUE = (0, 120, 212) # Accent blue
ACE_GRAY = (80, 80, 80)        # Body text
ACE_LIGHT_GRAY = (240, 240, 240)  # Table row alt
ACE_WHITE = (255, 255, 255)
ACE_BLACK = (20, 20, 20)
ACE_GREEN = (34, 139, 34)      # Totals highlight


def generate_proposal_pdf(
    session_data: Dict[str, Any],
    line_items: List[Dict[str, Any]],
    proposal_text: str,
    output_path: str,
) -> str:
    """
    Generate a professional PDF proposal.
    Returns the path to the generated PDF.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        story = []
        styles = getSampleStyleSheet()

        # ── Custom Styles ──────────────────────────────────────────────────
        title_style = ParagraphStyle(
            "BidTitle",
            parent=styles["Title"],
            fontSize=22,
            textColor=colors.Color(*[c/255 for c in ACE_BLUE]),
            spaceAfter=4,
            fontName="Helvetica-Bold",
            alignment=TA_LEFT,
        )
        subtitle_style = ParagraphStyle(
            "BidSubtitle",
            parent=styles["Normal"],
            fontSize=11,
            textColor=colors.Color(*[c/255 for c in ACE_LIGHT_BLUE]),
            spaceAfter=2,
            fontName="Helvetica",
        )
        header_style = ParagraphStyle(
            "SectionHeader",
            parent=styles["Heading2"],
            fontSize=12,
            textColor=colors.Color(*[c/255 for c in ACE_BLUE]),
            spaceBefore=14,
            spaceAfter=6,
            fontName="Helvetica-Bold",
            borderPad=2,
        )
        body_style = ParagraphStyle(
            "BidBody",
            parent=styles["Normal"],
            fontSize=9.5,
            textColor=colors.Color(*[c/255 for c in ACE_GRAY]),
            spaceAfter=6,
            leading=14,
            fontName="Helvetica",
            alignment=TA_JUSTIFY,
        )
        small_style = ParagraphStyle(
            "Small",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.gray,
            fontName="Helvetica",
        )
        total_style = ParagraphStyle(
            "TotalStyle",
            parent=styles["Normal"],
            fontSize=11,
            textColor=colors.Color(*[c/255 for c in ACE_BLUE]),
            fontName="Helvetica-Bold",
            alignment=TA_RIGHT,
        )

        # ── HEADER BLOCK ──────────────────────────────────────────────────
        company_name = session_data.get("company_name", "Ace Plumbing")
        company_address = session_data.get("company_address", "")
        company_phone = session_data.get("company_phone", "")
        company_email = session_data.get("company_email", "")
        company_license = session_data.get("company_license", "")

        # Header table: Company info left, Proposal info right
        proposal_num = f"BID-{datetime.now().strftime('%Y%m%d-%H%M')}"
        date_str = datetime.now().strftime("%B %d, %Y")

        header_left = [
            [Paragraph(f"<b>{company_name}</b>", ParagraphStyle("CH", fontSize=16, textColor=colors.Color(*[c/255 for c in ACE_BLUE]), fontName="Helvetica-Bold"))],
            [Paragraph(company_address or "Licensed & Insured Plumbing Contractor", small_style)],
            [Paragraph(company_phone or "", small_style)],
            [Paragraph(company_email or "", small_style)],
            [Paragraph(company_license or "", small_style)],
        ]

        header_right = [
            [Paragraph("<b>BID PROPOSAL</b>", ParagraphStyle("PR", fontSize=14, textColor=colors.Color(*[c/255 for c in ACE_BLUE]), fontName="Helvetica-Bold", alignment=TA_RIGHT))],
            [Paragraph(f"Proposal #: {proposal_num}", ParagraphStyle("PR2", fontSize=9, fontName="Helvetica", alignment=TA_RIGHT))],
            [Paragraph(f"Date: {date_str}", ParagraphStyle("PR3", fontSize=9, fontName="Helvetica", alignment=TA_RIGHT))],
            [Paragraph(f"Valid for: 30 days", ParagraphStyle("PR4", fontSize=9, fontName="Helvetica", alignment=TA_RIGHT))],
        ]

        header_table = Table(
            [[
                Table(header_left, colWidths=[3.5 * inch]),
                Table(header_right, colWidths=[3.5 * inch]),
            ]],
            colWidths=[4 * inch, 3.5 * inch],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(HRFlowable(width="100%", thickness=2, color=colors.Color(*[c/255 for c in ACE_BLUE]), spaceAfter=12))

        # ── CLIENT INFO ───────────────────────────────────────────────────
        client_name = session_data.get("client_name", "Valued Client")
        project_name = session_data.get("project_name", "Plumbing Project")
        project_type = session_data.get("project_type", "Residential").title()
        trade_type = session_data.get("trade_type", "Plumbing")

        client_data = [
            ["PREPARED FOR:", "PROJECT DETAILS:"],
            [client_name, project_name],
            ["", f"Type: {project_type} {trade_type}"],
            ["", f"Timeline: {session_data.get('timeline_notes', 'TBD')}"],
        ]
        client_table = Table(client_data, colWidths=[3.75 * inch, 3.75 * inch])
        client_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.Color(*[c/255 for c in ACE_LIGHT_BLUE])),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.Color(*[c/255 for c in ACE_BLACK])),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(client_table)
        story.append(Spacer(1, 0.15 * inch))

        # ── PROJECT SUMMARY ───────────────────────────────────────────────
        if session_data.get("project_summary"):
            story.append(Paragraph("PROJECT OVERVIEW", header_style))
            story.append(Paragraph(session_data["project_summary"], body_style))

        # ── PROPOSAL NARRATIVE ────────────────────────────────────────────
        if proposal_text:
            story.append(Paragraph("PROPOSAL", header_style))
            for para in proposal_text.split("\n\n"):
                para = para.strip()
                if para:
                    if para.isupper() or (para.endswith(":") and len(para) < 50):
                        story.append(Paragraph(para, ParagraphStyle(
                            "SubHead", fontSize=10, fontName="Helvetica-Bold",
                            textColor=colors.Color(*[c/255 for c in ACE_GRAY]),
                            spaceBefore=8, spaceAfter=4
                        )))
                    else:
                        story.append(Paragraph(para, body_style))

        # ── LINE ITEMS TABLE ──────────────────────────────────────────────
        story.append(Paragraph("DETAILED SCOPE & PRICING", header_style))

        table_data = [["#", "Description", "Qty", "Unit", "Material", "Labor", "Total"]]
        for i, item in enumerate(line_items, 1):
            est = "~" if item.get("is_estimated", True) else ""
            table_data.append([
                str(i),
                f"{est}{item.get('description', '')}",
                f"{float(item.get('quantity', 1)):.1f}",
                item.get("unit", "each"),
                f"${float(item.get('material_cost', 0)):,.2f}",
                f"${float(item.get('labor_cost', 0)):,.2f}",
                f"${float(item.get('total', 0)):,.2f}",
            ])

        col_widths = [0.3*inch, 2.8*inch, 0.5*inch, 0.8*inch, 0.9*inch, 0.8*inch, 0.9*inch]
        items_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        table_style = TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(*[c/255 for c in ACE_BLUE])),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            # Data rows
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.Color(*[c/255 for c in ACE_GRAY])),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (2, 1), (2, -1), "CENTER"),
            ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            # Alternating rows
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(*[c/255 for c in ACE_LIGHT_GRAY])]),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.25, colors.Color(0.8, 0.8, 0.8)),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.Color(*[c/255 for c in ACE_BLUE])),
            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ])
        items_table.setStyle(table_style)
        story.append(items_table)

        story.append(Paragraph(
            "<i>~ Estimated value based on document analysis. [CONFIRMED] = Client-provided value.</i>",
            small_style
        ))
        story.append(Spacer(1, 0.1 * inch))

        # ── TOTALS TABLE ──────────────────────────────────────────────────
        total_material = sum(float(i.get("material_cost", 0)) for i in line_items)
        total_labor = sum(float(i.get("labor_cost", 0)) for i in line_items)
        total_markup = sum(float(i.get("markup_amount", 0)) for i in line_items)
        grand_total = sum(float(i.get("total", 0)) for i in line_items)
        suggested_bid = grand_total * 1.05

        totals_data = [
            ["", "Material Costs:", f"${total_material:,.2f}"],
            ["", "Labor Costs:", f"${total_labor:,.2f}"],
            ["", "Markup / Overhead / Profit:", f"${total_markup:,.2f}"],
            ["", "TOTAL PROJECT COST:", f"${grand_total:,.2f}"],
            ["", "SUGGESTED BID PRICE:", f"${suggested_bid:,.2f}"],
        ]
        totals_table = Table(totals_data, colWidths=[4.5*inch, 1.8*inch, 1.2*inch])
        totals_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 2), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (1, 3), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (1, 3), (-1, -1), 11),
            ("TEXTCOLOR", (1, 3), (-1, 3), colors.Color(*[c/255 for c in ACE_BLUE])),
            ("TEXTCOLOR", (1, 4), (-1, 4), colors.Color(*[c/255 for c in ACE_GREEN])),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("LINEABOVE", (1, 3), (-1, 3), 1, colors.Color(*[c/255 for c in ACE_BLUE])),
            ("LINEABOVE", (1, 4), (-1, 4), 1, colors.Color(*[c/255 for c in ACE_GREEN])),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 0.2 * inch))

        # ── TERMS & CONDITIONS ────────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=8))
        story.append(Paragraph("TERMS & CONDITIONS", header_style))

        terms = session_data.get("terms_and_conditions") or """
**Payment Schedule:** 30% deposit required to schedule work. 40% due upon completion of rough-in and passing inspection. 30% balance due upon final completion and client walkthrough.

**Change Orders:** Any changes to the scope of work must be agreed upon in writing prior to execution. Additional work will be billed at current labor rates plus materials.

**Warranty:** All labor is warranted for one (1) year from date of substantial completion. All materials carry the manufacturer's standard warranty. Warranty is void if work is modified by others.

**Permits & Inspections:** Permit fees are included in this proposal unless otherwise noted. Contractor will schedule all required inspections.

**Liability:** Contractor carries general liability insurance and workers' compensation. Certificate of insurance available upon request.

**Acceptance:** This proposal is valid for 30 days from the date issued. Signing below constitutes acceptance of all terms and conditions.
        """.strip()

        for para in terms.split("\n\n"):
            para = para.strip()
            if para:
                # Convert **bold** markers
                para = para.replace("**", "<b>", 1).replace("**", "</b>", 1)
                story.append(Paragraph(para, small_style))
                story.append(Spacer(1, 3))

        # ── SIGNATURE BLOCK ───────────────────────────────────────────────
        story.append(Spacer(1, 0.3 * inch))
        sig_data = [
            ["CONTRACTOR ACCEPTANCE:", "", "CLIENT ACCEPTANCE:"],
            ["", "", ""],
            ["_" * 35, "", "_" * 35],
            [f"{company_name}", "", client_name],
            ["Authorized Signature / Date", "", "Authorized Signature / Date"],
        ]
        sig_table = Table(sig_data, colWidths=[2.8*inch, 1.9*inch, 2.8*inch])
        sig_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.Color(*[c/255 for c in ACE_BLUE])),
            ("TEXTCOLOR", (0, 3), (-1, 3), colors.Color(*[c/255 for c in ACE_GRAY])),
            ("TEXTCOLOR", (0, 4), (-1, 4), colors.grey),
            ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(sig_table)

        # ── FOOTER ────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.2 * inch))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=4))
        footer_text = f"{company_name} | {company_phone} | {company_email} | {company_license} | Generated by BidGenie AI"
        story.append(Paragraph(footer_text, ParagraphStyle(
            "Footer", fontSize=7, textColor=colors.grey, alignment=TA_CENTER, fontName="Helvetica"
        )))

        doc.build(story)
        logger.info(f"PDF generated: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        raise


def generate_csv_export(
    line_items: List[Dict[str, Any]],
    session_data: Dict[str, Any],
    output_path: str,
) -> str:
    """Generate CSV/Excel export of line items."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bid Proposal"

        # Header info
        blue_fill = PatternFill(start_color="004785", end_color="004785", fill_type="solid")
        light_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        ws["A1"] = session_data.get("company_name", "Ace Plumbing")
        ws["A1"].font = Font(bold=True, size=14, color="004785")
        ws["A2"] = f"Bid Proposal — {session_data.get('project_name', 'Project')}"
        ws["A2"].font = Font(size=11)
        ws["A3"] = f"Client: {session_data.get('client_name', '')} | Date: {datetime.now().strftime('%B %d, %Y')}"
        ws["A3"].font = Font(size=9, color="666666")

        ws.append([])

        # Column headers
        headers = ["#", "Description", "Qty", "Unit", "Material Cost", "Labor Cost", "Markup/OH/Profit", "Total", "Est?", "Notes"]
        ws.append(headers)
        header_row = ws.max_row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = blue_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for i, item in enumerate(line_items, 1):
            row = [
                i,
                item.get("description", ""),
                float(item.get("quantity", 1)),
                item.get("unit", "each"),
                float(item.get("material_cost", 0)),
                float(item.get("labor_cost", 0)),
                float(item.get("markup_amount", 0)),
                float(item.get("total", 0)),
                "Yes" if item.get("is_estimated", True) else "No",
                item.get("notes", ""),
            ]
            ws.append(row)
            data_row = ws.max_row
            fill = alt_fill if i % 2 == 0 else PatternFill()
            for col in range(1, 11):
                cell = ws.cell(row=data_row, column=col)
                if i % 2 == 0:
                    cell.fill = alt_fill
                if col in [5, 6, 7, 8]:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Totals
        ws.append([])
        total_material = sum(float(i.get("material_cost", 0)) for i in line_items)
        total_labor = sum(float(i.get("labor_cost", 0)) for i in line_items)
        total_markup = sum(float(i.get("markup_amount", 0)) for i in line_items)
        grand_total = sum(float(i.get("total", 0)) for i in line_items)
        suggested_bid = grand_total * 1.05

        for label, value, col in [
            ("Total Material", total_material, 5),
            ("Total Labor", total_labor, 6),
            ("Total Markup/OH/Profit", total_markup, 7),
            ("GRAND TOTAL", grand_total, 8),
            ("SUGGESTED BID", suggested_bid, 8),
        ]:
            ws.append(["", "", "", label, "", "", "", value])
            r = ws.max_row
            ws.cell(r, 4).font = Font(bold=True, size=9)
            ws.cell(r, 8).number_format = '"$"#,##0.00'
            ws.cell(r, 8).font = Font(bold=True, color="004785")
            ws.cell(r, 8).alignment = Alignment(horizontal="right")

        # Column widths
        col_widths = [4, 40, 6, 12, 14, 12, 18, 14, 6, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)
        logger.info(f"Excel generated: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel generation error: {e}")
        # Fallback to CSV
        import csv
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["#", "Description", "Qty", "Unit", "Material", "Labor", "Markup", "Total", "Estimated", "Notes"])
            for i, item in enumerate(line_items, 1):
                writer.writerow([
                    i, item.get("description", ""), item.get("quantity", 1),
                    item.get("unit", "each"), item.get("material_cost", 0),
                    item.get("labor_cost", 0), item.get("markup_amount", 0),
                    item.get("total", 0), item.get("is_estimated", True), item.get("notes", "")
                ])
        return csv_path
